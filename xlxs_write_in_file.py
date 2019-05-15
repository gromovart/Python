from openpyxl import Workbook
wb = Workbook()
# grab the active worksheet
ws = wb.active #Активация нужного листа
ws.title = "Лист1" #Название активного листа. Выбираем любое.
ws2 = wb.create_sheet("Лист2") #Создание Лист2
ws3 = wb.create_sheet("Лист3") #Создание Лист3
my_list = [['158-164'], ['170-176'], ['182-188'], ['44-46'], ['48-50'], ['52-54'], ['56-58'], ['60-62'], ['64-66']]
#Создаем таблицу 100х100 ячеек со значением 10

def write_xlsx(col,row,range_num):
  for x in range(range_num):
   for subarray in my_list:
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1
write_xlsx(2,2,2)
wb.save("sample.xlsx")