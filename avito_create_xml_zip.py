from xml.dom import minidom
from openpyxl import load_workbook
#подключаем библиотеку os
import os
from datetime import datetime
#Функия копирования файлов из одной папки в другую по списку файлов
import shutil
#текущая директория
cur_dir = str(os.getcwd())
#копирую дерево
b = []
not_found =[]

wb = load_workbook('файл_загрузки.xlsx')
ws = wb.active
id_list = []
title_list = []
desc_list = []
price_list = []
images_list = []
list_tag_variables = ['Id','Title','Description','Price','Images']
list_tag_permanent = ['DateBegin','DateEnd','AdStatus','AllowEmail','ManagerName','ContactPhone','Address','Category','GoodsType','AdType','Apparel','ListingFee']
temporary_list = [] #Список с именами файлов изображений
list_value_permanent = []

#Создаем объект
doc = minidom.Document()
#корневой тег Ads
ads = doc.createElement('Ads')
ads.setAttribute('formatVersion', '3')
ads.setAttribute('target', 'Avito.ru')
doc.appendChild(ads)

#Функция сохранения изображений в zip архив
def save_img_zip(add_quantity, num_str):
  time_now = datetime.now().strftime('%Y_%m_%d %H_%M_%S')
  #Создаем директорию для изображений
  folder = 'img_temp_' + time_now
  os.makedirs(folder)
  path = str(os.path.abspath(folder))
  #Добавляем изображения во временную папку
  for elm in temporary_list:
    try:
      shutil.copyfile('D:/Project/avito_xml/images_avito/' + elm, path + '/' + elm, follow_symlinks=True)
    except FileNotFoundError:
      not_found.append(elm)
  cut = int(add_quantity) + int(num_str)
  cut = str(cut)
  #Создаем zip архив
  shutil.make_archive(cur_dir + '/' + 'img_' + 'add_num_' + add_quantity + '_' + num_str + '_' + cut + '_' + time_now, 'zip', path)
  #Осматрваем директорию временной папки и сохраняем в переменную
  dir_list = os.listdir(path)
  #Рекурсивно удаляем файлы из временной директории
  for elm in dir_list:
  	try:
  	  os.remove(path + '/' + elm)
  	except FileNotFoundError:
  		break
  #Удаляем временную директорию		
  os.rmdir(path)

#Сохраняем в файл filename.xml.

def save_xml(add_quantity, num_str):
  xml_str = doc.toprettyxml(indent="  ")
  data = xml_str.replace("&lt;", "<").replace("&gt;", ">") #Удаляем баг замены тегов <> на &lt; &gt;
  folder = 'XML_file'
  if not os.path.exists(folder):
    os.makedirs(folder)
  path = str(os.path.abspath(folder))
  cut = int(add_quantity) + int(num_str)
  cut = str(cut)
  time_now = datetime.now().strftime('%Y_%m_%d %H_%M_%S')
  with open(path + '/' + 'add_num_' + add_quantity + '_' + num_str + '_' + cut + '_' + time_now + '.xml', "wb") as f:
    f.write(data.encode('utf-8'))
  f.close()

def check_quantity_add_file():
  check_list = []
  for i in range(2, 1500):
    check_list.append(str(ws.cell(row=i, column=1).value))
  for i in range(check_list.count('None')):
    check_list.remove('None')
  check_quantity = len(check_list)
  check_list.clear()
  return check_quantity

def add_time():
  list_add_vlue = ['', 'Да', 'Артем', '+ 7 (831) 238-93-08', 'Россия, Нижегородская область, Нижний Новгород, Торфяная улица, 9а', 'Одежда, обувь, аксессуары', 'Мужская одежда', 'Товар приобретен на продажу', 'Другое', 'Package']
  data_begin = input('Настройте время размещения объявлений\nУкажите дату строго в формате 2018-12-18T06:00:00+03:00\n-------------------------------------------------\nДАТА НАЧАЛА ПУБЛИКАЦИИ:')
  data_end = input('\n-------------------------------------------------\nДАТА СНЯТИЯ С ПУБЛИКАЦИИ:')
  list_value_permanent.append(data_begin)
  list_value_permanent.append(data_end)
  list_value_permanent.extend(list_add_vlue)
print('-------------------------------------------------')





#Функция чтения из файла загрузка.xlsx и заполнения списка списка.
def create_list(name_list, col_val):
  for i in range(2, 50):
    name_list.append(str(ws.cell(row=i, column=col_val).value))

#Заполняем списки данными из загрузка.xlsx.
def create_list_write():
  create_list(id_list, 1)
  create_list(title_list, 2)
  create_list(desc_list, 3)
  create_list(price_list, 4)
  create_list(images_list, 5)


create_list_write()




def create_tags_dev_2(test):
  ad = doc.createElement('Ad')
  for i in range(1):
    for l in range(len(list_tag_permanent)):
      tags_name = doc.createElement(list_tag_permanent[i + l])
      ad.appendChild(tags_name)
      for i in range(1):
        value = doc.createTextNode(list_value_permanent[i + l])
        tags_name.appendChild(value)

#Тег <Id>-----------------------------------------------------------------
#Выбираем нужный тег 
  for j in range(1):
    b = 0
    b = b + j
    tag_var_name = doc.createElement(list_tag_variables[b])
  ad.appendChild(tag_var_name)
  
#Подбираем нужное значение
  for elm in range(test):
    a = 0
    a = a + elm
    value_var = doc.createTextNode(id_list[a])

  tag_var_name.appendChild(value_var)
  
  #Тег <Title>--------------------------------------------------------------

  for j in range(2):
    b = 0
    b = b + j
    tag_var_name = doc.createElement(list_tag_variables[b])
  ad.appendChild(tag_var_name)
  
#Подбираем нужное значение
  for elm in range(test):
    a = 0
    a = a + elm
    value_var = doc.createTextNode(title_list[a])

  tag_var_name.appendChild(value_var)

#Тег <Description>---------------------------------------------------------

  for j in range(3):
    b = 0
    b = b + j
    tag_var_name = doc.createElement(list_tag_variables[b])
  ad.appendChild(tag_var_name)
  
#Подбираем нужное значение
  for elm in range(test):
    a = 0
    a = a + elm
    value_var = doc.createTextNode(desc_list[a])

  tag_var_name.appendChild(value_var)

#Тег <Price>-----------------------------------------------------------------

  for j in range(4):
    b = 0
    b = b + j
    tag_var_name = doc.createElement(list_tag_variables[b])
  ad.appendChild(tag_var_name)
  
#Подбираем нужное значение
  for elm in range(test):
    a = 0
    a = a + elm
    value_var = doc.createTextNode(price_list[a])

  tag_var_name.appendChild(value_var)
  
  #Тег <Images>------------------------------------------------------------

  for j in range(5):
    b = 0
    b = b + j
    tag_var_name = doc.createElement(list_tag_variables[b])
  tag_image = doc.createElement('Image')
  tag_var_name.appendChild(tag_image)

  ad.appendChild(tag_var_name)
  
#Подбираем нужное значение
  for elm in range(test):
    a = 0
    a = a + elm
    value_var = doc.createTextNode(images_list[a])
  temporary_list.append(images_list[a]) #Сохраняем значения имена файлов изображений во временный лист
  tag_image.setAttribute('name', images_list[a])

  ads.appendChild(ad)
  return ad

def main_fun(num_str, add_quantity):
  for i in range(add_quantity):
    x = num_str
    create_tags_dev_2(x + i)
  #file_mame_xml = input('Введите название файла для сохранения\nНазвание файла:')
  save_xml(str(add_quantity), str(num_str))
  print('-------------------------------------------------')
  save_img_zip(str(add_quantity), str(num_str))
  print('ФАЙЛ УСПЕШНО СОХРАНЕН!')



def main():
  add_time()
  print('-------------------------------------------------')
  print('Количество объявлений в файле: ' + str(check_quantity_add_file()))
  print('-------------------------------------------------')
  num_str = input('Укажите с какой строки начать добавление\nНОМЕР СТРОКИ:')
  add_quantity = input('Укажите количество объявлений для добавления в файл XML\nКОЛИЧЕСТВО:')
  print('-------------------------------------------------')
  main_fun(int(num_str), int(add_quantity))
  print('-------------------------------------------------')


if __name__ == '__main__':
    main()