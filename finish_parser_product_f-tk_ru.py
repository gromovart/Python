#Парсер товаров по списку категорий на сайте http://www.planeta-sirius.ru
#подключаем библиотеку os
import os
#подключаем бибилиотеку request
import requests
#подключаем бибилиотеку BeautifulSoup
from bs4 import BeautifulSoup
#подключаем бибилиотеку csv
from openpyxl import Workbook
wb = Workbook()
# grab the active worksheet
ws = wb.active #Активация нужного листа
ws.title = "Лист1" #Название активного листа. Выбираем любое.
ws2 = wb.create_sheet("Лист2") #Создание Лист2
ws3 = wb.create_sheet("Лист3") #Создание Лист3


#'https://www.f-tk.ru/catalog/01_spetsodezhda/01_01_spetsodezhda_zima/01_01_1_spetsodezhda_zima_kostyumy/' 
url = input('Введите ссылку на категорию\nСсылка:')
urls = [] #список ссылок на товары

def urls_list(url): 
  r = requests.get(url)
  soup = BeautifulSoup(r.text, 'html.parser')
  for i in soup.findAll('div', class_ = 'product-photo'):
    soup_a = i
    for j in soup_a.findAll('a', href = True):
      soup_b = j['href']
      urls.append('https://www.f-tk.ru' + soup_b)

def pars_articul(url):
  articul_list = []
  for url in urls:
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')
    for b in soup.find_all(class_ = 'product-name-link'):
      soup_a = b
      for b in range(1):
        articul = soup_a.find_all('span')[b].get_text()
        articul = articul.replace("Артикул:", "")
        articul = articul.strip()
        articul_list.append([articul])

  return(articul_list)

def pars_price(url): #Парсим цены товаров
  price = []
  for url in urls:
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')
    for i in soup.findAll('span', attrs={"itemprop": "price"}):
      i = [i.getText()]
      price.append(i)
  return(price)

def pars_name_product(url): #Парсим наименование товаров
  product_name = []
  number = '100'
  page = requests.get(url)
  soup = BeautifulSoup(page.text, 'html.parser')

  try:
    for b in range(int(number)):
      product = soup.find_all('strong')[b].get_text()
      articul = soup.find_all('span')[b].get_text()
      product_name.append([product])
  except IndexError:
    print('УСПЕХ')
  return(product_name)

def pars_img_links(url): #Парсим ссылки на изображения
  img_link_list = []
  for url in urls:
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')
    for i in soup.findAll(class_='product-photo idcarousel'):
      soup_c = i
      #print(soup_c)
      for j in soup_c.findAll(attrs={"itemprop": "image"}, src = True):
        link_img = j['src']
        img_link_list.append(['https://www.f-tk.ru' + link_img])
  try:
    for d in img_link_list:
      img_link_list.remove('https://www.f-tk.ru')
  except:
    print('Лишние элементы удалены')
  return(img_link_list)

def write_xlsx_table_name(col,row,range_num): #Заполняем заголовки таблицы
  value_table = [['Артикул'], ['Наименование товара'], ['Цена розничная'], ['Ссылки на картинки']]
  for x in range(range_num):
   for subarray in value_table:
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 1
     row += 0

def write_xlsx_product_name(col,row,range_num): #Заполняем столбец с наименованием
  for x in range(range_num):
   for subarray in pars_name_product(url):
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

def write_xlsx_price(col,row,range_num): #Заполняем столбец с ценами
  for x in range(range_num):
   for subarray in pars_price(url):
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

def write_xlsx_articul(col,row,range_num): #Заполняем столбец с артикулом
  for x in range(range_num):
   for subarray in pars_articul(url):
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

def write_xlsx_img_links(col,row,range_num): #Заполняем столбец с ссылками на товары
  for x in range(range_num):
   for subarray in pars_img_links(url):
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

urls_list(url)
print('-------------------------------------------------')
print('Список ссылок получен!')
print('-------------------------------------------------')
pars_articul(url)
print('-------------------------------------------------')
print('Список артикулов получен!')
print('-------------------------------------------------')
pars_price(url)
print('-------------------------------------------------')
print('Список цен получен!')
print('-------------------------------------------------')
pars_name_product(url)
print('-------------------------------------------------')
print('Список наименований получен!')
print('-------------------------------------------------')
pars_img_links(url)
print('-------------------------------------------------')
print('Список ссылок на изображения получен!')
print('-------------------------------------------------')
write_xlsx_table_name(1,1,1)
write_xlsx_articul(1,2,1)
write_xlsx_product_name(2,2,1)
write_xlsx_price(3,2,1)
write_xlsx_img_links(4,2,1)
name = input('Введите название файла для сохранения\nНазвание файла:')
wb.save(name + '.xlsx')
print('-------------------------------------------------')
print('РАБОТА ПАРСЕРА ЗАВЕРШЕНА УСПЕШНО! ФАЙЛ СОХРАНЕН!')
print('-------------------------------------------------')
