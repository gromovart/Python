from openpyxl import Workbook
wb = Workbook()
ws = wb.active #Активация нужного листа
ws.title = "ProductOptions" #Название активного листа. Выбираем любое.
ws2 = wb.create_sheet("ProductOptionValues") #Создание Лист2
ws3 = wb.create_sheet("Razmer_obuv") #Создание Лист3


product_id = [] #product_id
Product_Options = []
product_id_ProductOptions = [] #product_id лист ProductOptions
product_id_ProductOptionValues = [] #product_id лист Product_Option_Values
Product_Options_option = [['Ростовка'], ['Размер']] #option ProductOptions
Product_Options_value_option = [['Ростовка'], ['Ростовка'], ['Ростовка'], ['Размер'], ['Размер'], ['Размер'], ['Размер'], ['Размер'], ['Размер']] #option ProductOptions
Product_Options_option_new = [] #option ProductOptions
option_value = [] #option_value лист Product_Option_Values
razmer_obuv_id_product = []
razmer_obuv_id_product_option = [['38'], ['39'], ['40'], ['41'], ['42'], ['43'], ['44'], ['45'], ['46'], ['47']]
option = [['158-164'], ['170-176'], ['182-188'], ['44-46'], ['48-50'], ['52-54'], ['56-58'], ['60-62'], ['64-66']] #option_value лист Product_Option_Values

#Генерируем значения для столбца product_id
for i in range(50,746):
  product_id.append([i])

#Генерируем значения для столбца product_id лист ProductOptions
for m in product_id:
  product_id_ProductOptions.extend(([m,m]))

#Генерируем значения для столбца option лист ProductOptions
for b in range(746):
  Product_Options_option_new.extend(Product_Options_option)

#Генерируем значения для столбца product_id лист Product_Option_Values
for j in product_id:
  product_id_ProductOptionValues.extend([j, j, j, j, j, j,j,j,j])

#Генерируем значения для столбца option_value лист Product_Option_Values
for k in range(746):
  option_value.extend(option)

#Генерируем диапазон Razmer_obuv
for v in range(658,746):
  razmer_obuv_id_product.append([v])

#Генерируем значения для Razmer_obuv
for j in product_id:
  product_id_ProductOptionValues.extend([j, j, j, j, j, j,j,j,j])

#Функция записи product_id в файл xlsx
def write_product_id(col,row,range_num):
  for x in range(range_num):
   for subarray in product_id:
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

#write_product_id(1,2,1)

#Функция записи product_id_ProductOptions в файл xlsx
def write_product_id_ProductOptions(col,row,range_num):
  for x in range(range_num):
   for subarray in product_id_ProductOptions:
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

#write_xlsx(1,2,1)

#Функция записи Product_Options_option в файл xlsx
def write_Product_Options_option(col,row,range_num):
  for x in range(range_num):
   for subarray in Product_Options_option:
     for index, value in enumerate(subarray):
         ws.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

#write_Product_Options_option(1,2,1)

#Функция записи product_id_ProductOptionValues в файл xlsx
def write_product_id_ProductOptionValues(col,row,range_num):
  for x in range(range_num):
   for subarray in product_id_ProductOptionValues:
     for index, value in enumerate(subarray):
         ws2.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

#write_product_id_ProductOptionValues(1,2,1)

#Функция записи Product_Options_value_option в файл xlsx
def write_Product_Options_value_option(col,row,range_num):
  for x in range(range_num):
   for subarray in Product_Options_value_option:
     for index, value in enumerate(subarray):
         ws2.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

#write_Product_Options_value_option(1,2,695)

#Функция записи product_id_ProductOptionValues в файл xlsx
def write_Product_Options_value_option_value(col,row,range_num):
  for x in range(range_num):
   for subarray in option:
     for index, value in enumerate(subarray):
         ws2.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

#write_Product_Options_value_option_value(1,2,696)

def write_razmer_obuv_id_product(col,row,range_num):
  for x in range(range_num):
   for subarray in product_id_ProductOptionValues:
     for index, value in enumerate(subarray):
         ws3.cell(column=col+index, row=row).value = value
     col += 0
     row += 1

def main():
  write_product_id_ProductOptions(1,2,1)
  write_Product_Options_option(2,2,696)
  write_product_id_ProductOptionValues(1,2,1)
  write_Product_Options_value_option(2,2,696)
  write_Product_Options_value_option_value(3,2,696)
  write_razmer_obuv_id_product(1,2,1)
  wb.save("sample.xlsx")
#wb.save("sample.xlsx")

main()





